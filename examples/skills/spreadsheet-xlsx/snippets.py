"""Reusable openpyxl helpers for the spreadsheet-xlsx skill.

Copy this file next to your script (or paste the functions inline). Depends only
on `openpyxl` (`uv pip install openpyxl`).
"""

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_table(ws, rows, headers=None):
    """Write a list of dicts as a table with a bold, frozen header + autofilter.

    Cells keep their Python types, so numbers/dates stay numeric in Excel.
    `rows` is a list of dicts; `headers` defaults to the first row's keys.
    """
    if not rows:
        return
    headers = headers or list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    return headers


def _col_index(ws, header_name):
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column  # 1-based
    raise KeyError(f"column {header_name!r} not found in header row")


def apply_number_format(ws, header_name, fmt):
    """Apply an Excel number format to every data cell under `header_name`.

    Common formats: '"$"#,##0.00' (currency), '#,##0' (thousands),
    '0.0%' (percent — store 0.25 not 25), 'yyyy-mm-dd' (date).
    """
    col = _col_index(ws, header_name)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col).number_format = fmt


def autosize(ws, min_width=8, max_width=60, padding=2):
    """Size each column to the widest displayed value, clamped to [min,max]."""
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + padding))


def add_total(ws, header_name, label_col=1, label="Total"):
    """Append a totals row with a live =SUM() formula over the numeric column."""
    col = _col_index(ws, header_name)
    total_row = ws.max_row + 1
    letter = get_column_letter(col)
    ws.cell(row=total_row, column=label_col, value=label).font = Font(bold=True)
    cell = ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{ws.max_row})")
    cell.font = Font(bold=True)
    return total_row
